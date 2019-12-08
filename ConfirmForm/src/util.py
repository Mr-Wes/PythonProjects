import os
import openpyxl
import pandas as pd
import xml.dom.minidom
from shutil import copyfile
from PyQt5.QtWidgets import QMessageBox
from openpyxl.styles import Font


def readXML(xmlpath):
    # 打开xml文档
    dom = xml.dom.minidom.parse(xmlpath)
    root = dom.documentElement
    # 获取颜色列表
    color = getDataByTagName(root, 'color')
    # 获取型号列表
    type = getDataByTagName(root, 'type')
    # 获取主板型号
    board = getDataByTagName(root, 'board')
    # 获取显示屏型号
    screen = getDataByTagName(root, 'screen')
    # 获取摄像头型号
    camera = getDataByTagName(root, 'camera')
    # 获取二代证模块型号
    id = getDataByTagName(root, 'id')
    # 获取指纹仪型号
    finger = getDataByTagName(root, 'finger')
    # 获取电缆线型号
    wire = getDataByTagName(root, 'wire')
    return [color, type, board, screen, camera, id, finger, wire]


def getDataByTagName(root, tagname):
    result = []
    all = root.getElementsByTagName(tagname)[0]
    items = all.getElementsByTagName('item')
    for i in items:
        result.append(i.childNodes[0].data)
    return result


def testInput(widget, info):
    if info[0] == '':  # 检测订单号
        QMessageBox.information(widget, '提示', '请输入正确订单号！')
        return False
    elif info[1] == '':  # 检测客户名称
        QMessageBox.information(widget, '提示', '请输入正确客户名！')
        return False
    elif info[6] == '':  # 检测型号
        QMessageBox.information(widget, '提示', '请选择设备型号！')
        return False
    elif info[7] == '':  # 检测数量
        QMessageBox.information(widget, '提示', '请输入订单数量!')
        return False
    elif info[8] == '':  # 检查主板
        QMessageBox.information(widget, '提示', '请选择主板型号！')
        return False
    elif info[9] == '':  # 检测显示屏
        QMessageBox.information(widget, '提示', '请选择显示屏型号！')
        return False
    elif info[14] == '':  # 检测电缆线
        QMessageBox.information(widget, '提示', '请选择电缆线型号！')
        return False
    else:
        return True


def createFile(widget, source, filepath):
    # 检测文件是否存在
    if os.path.exists(filepath):
        QMessageBox.information(widget, '提示', '当前订单确认表已存在！')
    else:
        copyfile(source, filepath)
        return True


def writeFile(widget, info, filepath):
    df = pd.read_excel(filepath)
    df.iloc[0, 1] = info[0]  # 写入订单号
    df.iloc[2, 1] = info[1]  # 写入客户名称
    df.iloc[3, 1] = info[2]  # 写入商务
    df.iloc[3, 5] = info[3]  # 写入电话
    df.iloc[7, 3] = info[4]  # 写入颜色
    df.iloc[8, 3] = info[5]  # 写入丝印
    df.iloc[6, 3] = info[6]  # 写入型号
    df.iloc[0, 5] = info[7]  # 写入数量
    df.iloc[10, 3] = info[8]  # 写入主板
    df.iloc[12, 3] = info[9]  # 写入显示屏
    df.iloc[13, 3] = '无摄像头' if info[10] == '无' else info[10]  # 写入摄像头
    df.iloc[9, 3] = info[11]  # 写入键盘
    if info[12] != '无':
        df.iloc[14, 3] = info[12]  # 写入二代证
    if info[13] != '无':
        df.iloc[14, 4] = info[13]  # 写入指纹仪
    df.iloc[11, 3] = info[14]  # 写入电缆线
    df.iloc[18, 3] = info[15]  # 写入同步开机
    df.iloc[9, 3] += info[16]  # 写入国密
    df.iloc[15, 3] = info[17]  # 写入禾嘉读卡器
    df.iloc[15, 4] = info[18]  # 写入电池
    df.iloc[16, 3] = info[19]  # 写入4G
    df.iloc[17, 3] = info[20]  # 写入支架
    df.iloc[20, 2] = info[21]  # 写入软件版本
    df.iloc[19, 2] = info[22]  # 写入固件版本
    df.iloc[21, 2] = info[23]  # 写入M3版本
    df.iloc[26, 1] = info[24]  # 写入备注
    write = pd.ExcelWriter(filepath)
    df.to_excel(write, sheet_name='sheet1', index=False)
    write.close()
    # TODO 修整表格布局
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook.worksheets[0]
    worksheet.merge_cells('A1:F1')  # 合并第一行
    worksheet.merge_cells('B2:C2')  # 合并第二行
    worksheet.merge_cells('A3:F3')  # 合并第三行
    worksheet.merge_cells('B4:F4')  # 合并第四行
    worksheet.merge_cells('B5:C5')  # 合并第五行
    worksheet.merge_cells('A6:F6')  # 合并第六行
    worksheet.merge_cells('D7:F7')  # 合并第七行
    worksheet.merge_cells('D8:F8')  # 合并第八行
    worksheet.merge_cells('D9:F9')  # 合并第九行
    worksheet.merge_cells('D10:F10')  # 合并第十行
    worksheet.merge_cells('D11:F11')  # 合并第十一行
    worksheet.merge_cells('D12:F12')  # 合并第十二行
    worksheet.merge_cells('D13:F13')  # 合并第十三行
    worksheet.merge_cells('D14:F14')  # 合并第十四行
    worksheet.merge_cells('E15:F15')  # 合并第十五行
    worksheet.merge_cells('E16:F16')  # 合并第十六行
    worksheet.merge_cells('E17:F17')  # 合并第十七行
    worksheet.merge_cells('E18:F18')  # 合并第十八行
    worksheet.merge_cells('D19:F19')  # 合并第十九行
    worksheet.merge_cells('D20:F20')  # 合并第二十行
    worksheet.merge_cells('C21:F21')  # 合并第二十一行
    worksheet.merge_cells('C22:F22')  # 合并第二十二行
    worksheet.merge_cells('C23:F23')  # 合并第二十三行
    worksheet.merge_cells('C24:F24')  # 合并第二十四行
    worksheet.merge_cells('B25:F25')  # 合并第二十五行
    worksheet.merge_cells('B26:F26')  # 合并第二十六行
    worksheet.merge_cells('B27:F27')  # 合并第二十七行
    worksheet.merge_cells('B28:F28')  # 合并第二十八行
    # worksheet.merge_cells('C24:F24')  # 合并第二十九行
    worksheet.merge_cells('A8:A20')  # 合并第一列
    worksheet.merge_cells('B8:B20')  # 合并第二列
    worksheet.merge_cells('C15:C18')  # 合并第三列
    # 制定样式
    worksheet['A3'].font = Font('微软雅黑', size = 10)
    workbook.save(filepath)
    return True
